#!/Library/Frameworks/Python.framework/Versions/3.6/bin/python3
''' Visual pinout of a bga or connector '''
from re import split, match, findall
from sys import argv
from logging import basicConfig, getLogger, INFO
from json import load, dump
from os.path import join, basename
from os import getcwd
from argparse import ArgumentParser, RawDescriptionHelpFormatter
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from natsort import natsorted
from PyQt5.QtWidgets import QApplication, QGraphicsView, QGraphicsScene
from PyQt5.QtGui import QPainter, QBrush, QColor, QImage, QPixmap
from PyQt5.QtCore import QRectF, Qt, QPoint


class PartViewer(QGraphicsView):
    ''' Create a render of the part and load it into a QWidget '''
    def __init__(self, part, partview_settings):
        self.settings = partview_settings
        self.part = part
        self.image = None
        self.box_size = 50
        self.total_steps = 0
        super(PartViewer, self).__init__()
        self.generateRender()

    def initUI(self):
        ''' Init all the UI elements '''
        self.setWindowTitle(self.settings['title'])
        pixmap = QPixmap.fromImage(self.image)
        item = pixmap.scaled(self.settings['width'],
                             self.settings['height'],
                             Qt.KeepAspectRatio,
                             Qt.SmoothTransformation)
        self.scene = QGraphicsScene()
        self.scene.addPixmap(item)
        self.setScene(self.scene)
        self.setTransformationAnchor(self.AnchorUnderMouse)
        self.setResizeAnchor(self.AnchorUnderMouse)
        self.setDragMode(self.ScrollHandDrag)
        self.show()

    def resetZoom(self):
        ''' Go back to a scale of 1.0 '''
        self.fitInView(QRectF(0, 0, self.settings['width'], self.settings['height']))

    def wheelEvent(self, event):
        ''' If the wheel is scrolled; figure out how much to zoom '''
        steps = (event.angleDelta().y() / 120)
        self.total_steps += steps
        if (self.total_steps * steps) < 0:
            self.total_steps = steps  # Zoom out by steps
        factor = 1.0 + (steps / 10.0)
        self.settings['factor'] *= factor
        if self.settings['factor'] < .6:
            self.settings['factor'] = .6
        elif self.settings['factor'] > 5:
            self.settings['factor'] = 5
            return
        else:
            self.scale(factor, factor)

    def generateRender(self):
        ''' Generate the part '''
        part_cols = self.part.getColumns()
        part_rows = self.part.getRows()
        self.scaleBoxSize(part_cols, part_rows)
        self.image = QImage(self.settings['width'],
                            self.settings['height'] + (self.box_size / 2),
                            QImage.Format_RGB32)  # 2k image
        self.image.fill(QColor('#ffffff'))  # Background Color
        paint = QPainter(self.image)
        paint.setRenderHints(QPainter.HighQualityAntialiasing |
                             QPainter.SmoothPixmapTransform |
                             QPainter.TextAntialiasing, True)
        if self.settings['rotate']:
            paint.translate(QPoint(self.settings['width'], 0))
            paint.rotate(90)
        for hdr_offset, column in enumerate(part_cols):
            paint.drawText(QRectF(hdr_offset * self.box_size + (self.box_size / 2),
                                  0,
                                  self.box_size,
                                  self.box_size),
                           Qt.AlignCenter,
                           column)
        for y_offset, row in enumerate(part_rows):
            for x_offset, column in enumerate(part_cols):
                pin = self.part.getPin(str(row) + column)
                if pin and pin['name'] is not None:
                    fill = pin['color']
                    brush = QBrush(QColor(fill))
                    paint.setBrush(brush)
                    rect = QRectF(self.box_size * x_offset + (self.box_size / 2),
                                  self.box_size * y_offset + self.box_size,
                                  self.box_size,
                                  self.box_size)
                    if self.settings['circles']:
                        paint.drawEllipse(rect)
                    else:
                        paint.drawRect(rect)
                    paint.drawText(rect, Qt.AlignCenter, pin['name'][:6])
            paint.drawText(QRectF(self.box_size * len(part_cols) + (self.box_size / 2),
                                  self.box_size * y_offset + self.box_size,
                                  self.box_size,
                                  self.box_size),
                           Qt.AlignCenter,
                           row)
        paint.end()

    def save(self):
        ''' Save the Pixmap as a .png '''
        save_file = join(PATH, self.settings['title'] + '.png')
        print(f'Saved to {save_file}')
        self.image.save(save_file)

    def scaleBoxSize(self, columns, rows):
        ''' If the part width is greater than 1536 (2K width) scale up or down
        '''
        part_width = (len(columns) + 1) * self.box_size
        if part_width < 1536.0:
            window_scale = 1536.0/part_width
        else:
            window_scale = part_width/1536.0
        self.box_size = self.box_size * window_scale
        self.settings['width'] = (len(columns) + 1) * self.box_size + self.box_size
        self.settings['height'] = (len(rows) + 1) * self.box_size + self.box_size


class PartObject(object):
    ''' Load and create a part from a source '''
    def __init__(self, pins, filename):
        super(PartObject, self).__init__()
        self.__pins = pins
        self.__columns, self.__rows = self.sortAndSpiltPinList()
        self.filename = basename(filename).split('.')[0]

    @classmethod
    def fromExcel(cls, filename):
        ''' Import an Excel and create a PartObject '''
        number = 'Number'
        name = 'Name'
        workbook = load_workbook(filename)
        sheet = workbook.active  # Grab the first sheet
        header = sheet.iter_cols(max_row=1)
        try:
            column = getColIndexOfHeader([number, name], header)
            bga = dict()
            for excel_row in range(2, sheet.max_row + 1):
                pin = sheet.cell(row=excel_row, column=column[number]).value
                net = sheet.cell(row=excel_row, column=column[name])
                if pin is not None or net.value is not None:
                    if net.fill.patternType == 'solid':
                        bga.update({pin: {'name': net.value,
                                          'color': str('#' + net.fill.start_color.rgb[2:])}})
                    else:
                        bga.update({pin: {'name': net.value,
                                          'color': '#ffffff'}})
        except (TypeError, ValueError, KeyError, UnboundLocalError) as error:
            print(error)
            raise
        return cls(bga, filename)

    @classmethod
    def fromTelesis(cls, filename, refdes):
        ''' Import a Telesis formatted file and create a PartObject '''
        with open(filename, 'r') as tel_file:
            tel_text = tel_file.readlines()
        tel_netlist = dict()
        for line in tel_text:
            reg = match(r'(.*);', line)
            reg2 = findall(refdes + r'\.([a-zA-Z0-9]+)', line)
            if reg and reg2:
                net = reg.group(1)
                for reg_match in reg2:
                    pin = reg_match
                    tel_netlist.update(
                        {pin: {'name': net, 'color': '#ffffff'}})
        return cls(tel_netlist, filename)

    @classmethod
    def fromJson(cls, filename):
        ''' Import a json file with a format {pin: {name:, color:}} '''
        return cls(load(open(filename)), filename)

    def addPin(self, pin, net, color):
        ''' Add a new pin to the part '''
        self.__pins.update({pin: {'name': net, 'color': color}})

    def getColumns(self):
        ''' Get the columns in a part. [A - AZ] '''
        return self.__columns

    def getRows(self):
        ''' Get the rows in a part.  [1-n] '''
        return self.__rows

    def getPin(self, pin):
        ''' Get the name and color of a pin '''
        if pin in self.__pins:
            return self.__pins[pin]
        return None

    def getPins(self):
        ''' Return the pin names '''
        return self.__pins.keys()

    def getNumberofPins(self):
        ''' Return how many pins are in the part '''
        return len(self.__pins)

    def getNames(self):
        ''' Return the net names '''
        return self.__pins.values()['name']

    def dumpJson(self):
        ''' Dump the PartObject dictionary to a .json file '''
        save_file = join(PATH, self.filename + '.json')
        print(f'Saved to {save_file}')
        with open(save_file, 'w') as outfile:
            dump(self.__pins, outfile, indent=4, separators=(',', ': '))

    def sortAndSpiltPinList(self):
        ''' Take a list of pins and spilt by letter and number then sort '''
        r_list = list()
        c_list = list()
        for pin in self.getPins():
            split_pin = split(r'(\d+)', pin)
            if split_pin[0] not in r_list:
                r_list.append(split_pin[0])
            c_list.append(split_pin[1])
        temp = list()
        temp2 = list()
        for item in r_list:
            if len(item) > 1:
                temp.append(item)
            else:
                temp2.append(item)
        temp2 = natsorted(temp2)
        temp2.extend(natsorted(temp))
        return natsorted(set(c_list)), temp2


def getColIndexOfHeader(name, columns):
    ''' return a list of the column numbers if it matches '''
    indexes = dict()
    for col in columns:
        if col[0].value in name:  # Only look at the first row
            indexes.update(
                {col[0].value: column_index_from_string(col[0].column)})
    return indexes


def setupLogger(logger_filename):
    ''' Configure a logger at the current level= setting '''
    log_file = join(PATH, logger_filename)
    basicConfig(filename=log_file,
                format='%(levelname)s: %(message)s',
                filemode="w",
                level=INFO)
    logger = getLogger('root')
    return logger


def parseCommandLine():
    ''' Handle any command line input '''
    usage = '''USAGE:
    part_map.py -e part_spreadsheet.xlsx
    part_map.py -j part.json
    part_map.py -rsdt part.tel'''
    parser = ArgumentParser(formatter_class=RawDescriptionHelpFormatter,
                            epilog=usage)
    parser.add_argument('filename',
                        nargs='?',
                        default='artix7_example.xlsx',
                        help='The  file to read in')
    parser.add_argument('--excel', '-e',
                        action='store_true',
                        help='Load from an excel file')
    parser.add_argument('--json', '-j',
                        action='store_true',
                        help='Load from a json file')
    parser.add_argument('--tel', '-t',
                        action='store_true',
                        help='Load from a Telesis file')
    parser.add_argument('--refdes',
                        help='The refdes to pull from the Telesis')
    parser.add_argument('--circles', '-c',
                        action='store_true',
                        help='Draw using circles instead of rectangles')  
    parser.add_argument('--rotate', '-r',
                        action='store_true',
                        help='Rotate the image by 90 degrees')
    parser.add_argument('--save', '-s',
                        action='store_true',
                        help='Save the image as a .png')
    parser.add_argument('--dump', '-d',
                        action='store_true',
                        help='Dump PartObject as a Json File')
    parser.add_argument('--nogui', '-n',
                        action='store_true',
                        help='Do not open GUI window')
    return parser.parse_args()


def main():
    ''' Handle all the input and create the viewer '''
    args = parseCommandLine()
    if args.excel:
        part = PartObject.fromExcel(args.filename)
    elif args.json:
        part = PartObject.fromJson(args.filename)
    elif args.tel:
        part = PartObject.fromTelesis(args.filename, args.refdes)
    else:
        exit()
    global APP
    APP = QApplication(argv)
    screen_resolution = APP.desktop().screenGeometry()
    view_settings = {'width': screen_resolution.width(),
                     'height': screen_resolution.height(),
                     'title': basename(args.filename).split('.')[0],
                     'rotate': args.rotate,
                     'circles': args.circles,
                     'factor': 1.0}
    view = PartViewer(part, view_settings)
    if not args.nogui:
        view.initUI()
    if args.dump:
        part.dumpJson()
    if args.save:
        view.save()
    if args.nogui:
        APP.closeAllWindows()
    else:
        exit(APP.exec())


APP = None
PATH = getcwd()
# Setup Global Logger module
# LOG = setupLogger('part_map.txt')

if __name__ == '__main__':
    main()
